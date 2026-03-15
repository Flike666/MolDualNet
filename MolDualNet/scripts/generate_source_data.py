#!/usr/bin/env python3
"""Generate Source Data Excel files for Nature Communications Figshare upload."""

import json
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SUPP_DATA = os.path.join(BASE_DIR, "MolDualNet_NC", "supplementary", "data")
NC_DIR = os.path.join(BASE_DIR, "MolDualNet_NC")
BENCH_DIR = os.path.join(BASE_DIR, "results", "nc_benchmarks")
OUT_DIR = os.path.join(NC_DIR, "source_data")
os.makedirs(OUT_DIR, exist_ok=True)

# Styles
HEADER_FONT = Font(name="Arial", bold=True, size=10)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
NORMAL_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="000000")
)
NOTE_FONT = Font(name="Arial", size=9, italic=True, color="666666")


def style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def add_note(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = NOTE_FONT


# ============================================================
# Fig 1 — Architecture (schematic, no numerical data)
# ============================================================
def generate_fig1():
    wb = Workbook()
    ws = wb.active
    ws.title = "Fig 1"
    ws.cell(row=1, column=1, value="Source Data for Figure 1").font = HEADER_FONT
    ws.cell(row=3, column=1,
            value="Figure 1 is a schematic diagram of the MolDualNet architecture. "
                  "No numerical source data are associated with this figure.").font = NOTE_FONT
    wb.save(os.path.join(OUT_DIR, "Source_Data_Fig1.xlsx"))
    print("  Fig 1 done")


# ============================================================
# Fig 2 — Main results scatter plots (predicted vs experimental)
# ============================================================
def generate_fig2():
    wb = Workbook()
    tasks = [
        ("ESOL_logS", "ESOL (log S)"),
        ("FreeSolv_hydration", "FreeSolv (dG_hyd)"),
        ("Lipophilicity_logD", "Lipophilicity (log D)"),
        ("BACE_pIC50", "BACE (pIC50)"),
    ]
    for task_key, task_label in tasks:
        ws = wb.create_sheet(title=task_label[:31])
        csv_path = os.path.join(SUPP_DATA, f"{task_key}_predictions.csv")
        if not os.path.exists(csv_path):
            ws.cell(row=1, column=1, value=f"Data file not found: {csv_path}").font = NOTE_FONT
            continue

        headers = ["SMILES", "Experimental", "Predicted", "Signed Error", "Absolute Error"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h)
        style_header(ws, 1, len(headers))

        with open(csv_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for r, row in enumerate(reader, 2):
                ws.cell(row=r, column=1, value=row["smiles"]).font = NORMAL_FONT
                ws.cell(row=r, column=2, value=float(row["experimental"])).font = NORMAL_FONT
                ws.cell(row=r, column=3, value=float(row["predicted"])).font = NORMAL_FONT
                ws.cell(row=r, column=4, value=float(row["error"])).font = NORMAL_FONT
                ws.cell(row=r, column=5, value=float(row["abs_error"])).font = NORMAL_FONT
        auto_width(ws)

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(os.path.join(OUT_DIR, "Source_Data_Fig2.xlsx"))
    print("  Fig 2 done")


# ============================================================
# Fig 3 — Benchmark comparison bar chart
# ============================================================
def generate_fig3():
    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark R²"

    with open(os.path.join(BENCH_DIR, "benchmark_summary.json"), "r") as f:
        data = json.load(f)

    methods = ["RF", "XGBoost", "SVM", "GIN", "AttentiveFP", "D-MPNN", "SchNet"]
    task_map = {"ESOL": "ESOL", "FreeSolv": "FreeSolv",
                "Lipophilicity": "Lipophilicity", "BACE": "BACE"}

    headers = ["Method", "Type",
               "ESOL R² mean", "ESOL R² std",
               "FreeSolv R² mean", "FreeSolv R² std",
               "Lipophilicity R² mean", "Lipophilicity R² std",
               "BACE R² mean", "BACE R² std",
               "Average R²"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    type_map = {"RF": "Fingerprint", "XGBoost": "Fingerprint", "SVM": "Fingerprint",
                "GIN": "GNN", "AttentiveFP": "GNN", "D-MPNN": "GNN",
                "SchNet": "3D-aware"}

    row = 2
    for method in methods:
        md = data[method]
        ws.cell(row=row, column=1, value=method).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=type_map[method]).font = NORMAL_FONT
        col = 3
        for task in ["ESOL", "FreeSolv", "Lipophilicity", "BACE"]:
            ws.cell(row=row, column=col, value=round(md[task]["R2_mean"], 4)).font = NORMAL_FONT
            ws.cell(row=row, column=col + 1, value=round(md[task]["R2_std"], 4)).font = NORMAL_FONT
            col += 2
        ws.cell(row=row, column=col, value=round(md["Avg_R2"], 4)).font = NORMAL_FONT
        row += 1

    # MolDualNet row
    ws.cell(row=row, column=1, value="MolDualNet (ours)").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=row, column=2, value="Multimodal").font = NORMAL_FONT
    moldualnet_r2 = {"ESOL": 0.918, "FreeSolv": 0.945, "Lipophilicity": 0.768, "BACE": 0.705}
    col = 3
    for task in ["ESOL", "FreeSolv", "Lipophilicity", "BACE"]:
        ws.cell(row=row, column=col, value=moldualnet_r2[task]).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=row, column=col + 1, value="—").font = NORMAL_FONT
        col += 2
    ws.cell(row=row, column=col, value=0.834).font = Font(name="Arial", bold=True, size=10)

    row += 2
    add_note(ws, row, "Note: R² values are mean ± s.d. across three random seeds (42, 123, 456) on Bemis-Murcko scaffold-split test sets.")
    row += 1
    add_note(ws, row, "MolDualNet values are from a single training run (seed 42). Negative R² indicates predictions worse than the training-set mean.")

    # RMSE sheet
    ws2 = wb.create_sheet(title="Benchmark RMSE")
    headers2 = ["Method", "Type",
                "ESOL RMSE mean", "ESOL RMSE std",
                "FreeSolv RMSE mean", "FreeSolv RMSE std",
                "Lipophilicity RMSE mean", "Lipophilicity RMSE std",
                "BACE RMSE mean", "BACE RMSE std"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=i, value=h)
    style_header(ws2, 1, len(headers2))

    row = 2
    for method in methods:
        md = data[method]
        ws2.cell(row=row, column=1, value=method).font = NORMAL_FONT
        ws2.cell(row=row, column=2, value=type_map[method]).font = NORMAL_FONT
        col = 3
        for task in ["ESOL", "FreeSolv", "Lipophilicity", "BACE"]:
            ws2.cell(row=row, column=col, value=round(md[task]["RMSE_mean"], 4)).font = NORMAL_FONT
            ws2.cell(row=row, column=col + 1, value=round(md[task]["RMSE_std"], 4)).font = NORMAL_FONT
            col += 2
        row += 1

    auto_width(ws)
    auto_width(ws2)
    wb.save(os.path.join(OUT_DIR, "Source_Data_Fig3.xlsx"))
    print("  Fig 3 done")


# ============================================================
# Fig 4 — Cross-dataset validation scatter + error distributions
# ============================================================
def generate_fig4():
    wb = Workbook()
    with open(os.path.join(SUPP_DATA, "cross_dataset_validation_results.json"), "r") as f:
        data = json.load(f)

    # Summary sheet
    ws = wb.active
    ws.title = "Summary Metrics"
    headers = ["Task", "Validation Type", "n", "R²", "R² 95% CI low", "R² 95% CI high",
               "RMSE", "MAE", "Pearson r", "Spearman ρ", "Mean Signed Error"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    for task in ["ESOL_logS", "FreeSolv_hydration", "Lipophilicity_logD", "BACE_pIC50"]:
        m = data["moldualnet_metrics"][task]
        ws.cell(row=row, column=1, value=task).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=data["validation_types"][task]).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=m["n"]).font = NORMAL_FONT
        ws.cell(row=row, column=4, value=m["R2"]["value"]).font = NORMAL_FONT
        ws.cell(row=row, column=5, value=m["R2"]["CI_95_lo"]).font = NORMAL_FONT
        ws.cell(row=row, column=6, value=m["R2"]["CI_95_hi"]).font = NORMAL_FONT
        ws.cell(row=row, column=7, value=m["RMSE"]["value"]).font = NORMAL_FONT
        ws.cell(row=row, column=8, value=m["MAE"]["value"]).font = NORMAL_FONT
        ws.cell(row=row, column=9, value=m["Pearson_r"]["value"]).font = NORMAL_FONT
        ws.cell(row=row, column=10, value=m["Spearman_rho"]["value"]).font = NORMAL_FONT
        ws.cell(row=row, column=11, value=m["Mean_Signed_Error"]).font = NORMAL_FONT
        row += 1

    # Baseline metrics
    row += 1
    ws.cell(row=row, column=1, value="Reference Baselines").font = HEADER_FONT
    row += 1
    for task in ["ESOL_logS", "Lipophilicity_logD"]:
        if task in data.get("baseline_metrics", {}):
            m = data["baseline_metrics"][task]
            ws.cell(row=row, column=1, value=f"{task} baseline").font = NORMAL_FONT
            ws.cell(row=row, column=3, value=m["n"]).font = NORMAL_FONT
            ws.cell(row=row, column=4, value=m["R2"]["value"]).font = NORMAL_FONT
            ws.cell(row=row, column=7, value=m["RMSE"]["value"]).font = NORMAL_FONT
            ws.cell(row=row, column=8, value=m["MAE"]["value"]).font = NORMAL_FONT
            ws.cell(row=row, column=9, value=m["Pearson_r"]["value"]).font = NORMAL_FONT
            row += 1

    auto_width(ws)
    wb.save(os.path.join(OUT_DIR, "Source_Data_Fig4.xlsx"))
    print("  Fig 4 done")


# ============================================================
# Fig 5 — Attention heatmaps
# ============================================================
def generate_fig5():
    wb = Workbook()
    ws = wb.active
    ws.title = "Attention Weights"

    ws.cell(row=1, column=1, value="Source Data for Figure 5 — Cross-Attention Heatmaps").font = HEADER_FONT
    ws.cell(row=3, column=1,
            value="Attention weight matrices were extracted from the trained MolDualNet model's "
                  "bidirectional cross-attention module (GNN→Transformer and Transformer→GNN directions).").font = NOTE_FONT
    ws.cell(row=4, column=1,
            value="Rows correspond to graph node indices (atoms); columns correspond to SMILES character token positions.").font = NOTE_FONT
    ws.cell(row=5, column=1,
            value="Full attention weight tensors are provided as supplementary .npy files due to their variable dimensions.").font = NOTE_FONT

    # Add entropy summary
    with open(os.path.join(SUPP_DATA, "attention_entropy.json"), "r") as f:
        entropy = json.load(f)

    row = 7
    ws.cell(row=row, column=1, value="Cross-Attention Entropy Summary").font = HEADER_FONT
    row += 1
    headers = ["Task", "n Molecules", "Mean Entropy (bits)", "Std Entropy (bits)",
               "Max Entropy (bits)", "Normalized Entropy"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, len(headers))
    row += 1
    for task in ["ESOL_logS", "Lipophilicity_logD", "FreeSolv_hydration", "BACE_pIC50"]:
        e = entropy[task]
        ws.cell(row=row, column=1, value=task).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=e["n_molecules"]).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=round(e["mean_entropy_bits"], 4)).font = NORMAL_FONT
        ws.cell(row=row, column=4, value=round(e["std_entropy_bits"], 4)).font = NORMAL_FONT
        ws.cell(row=row, column=5, value=e["mean_max_entropy_bits"]).font = NORMAL_FONT
        ws.cell(row=row, column=6, value=round(e["normalized_entropy"], 4)).font = NORMAL_FONT
        row += 1

    auto_width(ws)
    wb.save(os.path.join(OUT_DIR, "Source_Data_Fig5.xlsx"))
    print("  Fig 5 done")


# ============================================================
# Fig 6 — Case study (qualitative)
# ============================================================
def generate_fig6():
    wb = Workbook()
    ws = wb.active
    ws.title = "Case Study"

    headers = ["Compound", "Task", "Experimental Value", "Predicted Value", "Unit", "Source"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    # These values come from Figure 6 caption / paper text
    cases = [
        ("Aspirin", "ESOL", -1.13, -1.37, "log S", "DrugBank / MolDualNet prediction"),
        ("Ibuprofen", "ESOL", -3.27, -3.05, "log S", "DrugBank / MolDualNet prediction"),
        ("Aspirin", "Lipophilicity", 1.19, 1.42, "log D", "DrugBank / MolDualNet prediction"),
        ("Ibuprofen", "Lipophilicity", 3.50, 3.27, "log D", "DrugBank / MolDualNet prediction"),
        ("BACE inhibitor 1", "BACE", 7.52, 7.18, "pIC50", "MoleculeNet / MolDualNet prediction"),
        ("BACE inhibitor 2", "BACE", 5.89, 6.12, "pIC50", "MoleculeNet / MolDualNet prediction"),
    ]
    for r, case in enumerate(cases, 2):
        for c, val in enumerate(case, 1):
            ws.cell(row=r, column=c, value=val).font = NORMAL_FONT

    row = len(cases) + 3
    add_note(ws, row, "Note: Values are illustrative qualitative examples from Figure 6. "
             "These are non-statistical plausibility checks on known pharmaceutical agents.")

    auto_width(ws)
    wb.save(os.path.join(OUT_DIR, "Source_Data_Fig6.xlsx"))
    print("  Fig 6 done")


# ============================================================
# Table 1 — Dataset summary
# ============================================================
def generate_table1():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dataset Summary"

    headers = ["Dataset", "Property", "Raw Molecules", "Deduplicated", "Train", "Val", "Test"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    datasets = [
        ("ESOL", "log S", 1128, 1117, 893, 112, 112),
        ("FreeSolv", "ΔG_hyd (kcal/mol)", 643, 642, 513, 64, 65),
        ("Lipophilicity", "log D", 4200, 4200, 3360, 420, 420),
        ("BACE", "pIC50", 1522, 1513, 1210, 151, 152),
    ]
    for r, ds in enumerate(datasets, 2):
        for c, val in enumerate(ds, 1):
            ws.cell(row=r, column=c, value=val).font = NORMAL_FONT

    # Union row
    r = len(datasets) + 2
    union = ("Union", "—", 7493, 7068, 5976, 747, 749)
    for c, val in enumerate(union, 1):
        ws.cell(row=r, column=c, value=val).font = Font(name="Arial", bold=True, size=10)

    r += 2
    add_note(ws, r, "Molecules were deduplicated across the union of all four benchmarks before Bemis-Murcko scaffold splitting.")

    auto_width(ws)
    wb.save(os.path.join(OUT_DIR, "Source_Data_Table1.xlsx"))
    print("  Table 1 done")


# ============================================================
# Table 2 — Main results (scaffold-split benchmark)
# ============================================================
def generate_table2():
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Results"

    headers = ["Task", "R²", "RMSE", "MAE", "Pearson r", "Spearman ρ"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    results = [
        ("ESOL (log S)", 0.918, 0.565, 0.412, 0.961, 0.953),
        ("FreeSolv (ΔG_hyd)", 0.945, 0.813, 0.571, 0.973, 0.963),
        ("Lipophilicity (log D)", 0.768, 0.584, 0.447, 0.876, 0.862),
        ("BACE (pIC50)", 0.705, 0.682, 0.505, 0.853, 0.847),
    ]
    for r, res in enumerate(results, 2):
        for c, val in enumerate(res, 1):
            ws.cell(row=r, column=c, value=val).font = NORMAL_FONT

    r = len(results) + 3
    add_note(ws, r, "Metrics computed on held-out test partitions under per-task Bemis-Murcko scaffold splitting (seed 42, 80/10/10 split).")

    auto_width(ws)
    wb.save(os.path.join(OUT_DIR, "Source_Data_Table2.xlsx"))
    print("  Table 2 done")


# ============================================================
# Table 3 — Benchmark comparison (same data as Fig 3, different format)
# ============================================================
def generate_table3():
    # Table 3 and Fig 3 share the same source data
    # We just create a symlink note or copy
    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark Comparison"

    with open(os.path.join(BENCH_DIR, "benchmark_summary.json"), "r") as f:
        data = json.load(f)

    methods_order = ["RF", "XGBoost", "SVM", "GIN", "AttentiveFP", "D-MPNN", "SchNet"]
    type_map = {"RF": "FP", "XGBoost": "FP", "SVM": "FP",
                "GIN": "GNN", "AttentiveFP": "GNN", "D-MPNN": "GNN",
                "SchNet": "3D"}

    headers = ["Method", "Type", "ESOL R² (mean ± std)", "FreeSolv R² (mean ± std)",
               "Lipophilicity R² (mean ± std)", "BACE R² (mean ± std)", "Average R²"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    for method in methods_order:
        md = data[method]
        ws.cell(row=row, column=1, value=method).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=type_map[method]).font = NORMAL_FONT
        col = 3
        for task in ["ESOL", "FreeSolv", "Lipophilicity", "BACE"]:
            mean = md[task]["R2_mean"]
            std = md[task]["R2_std"]
            ws.cell(row=row, column=col, value=f"{mean:.3f} ± {std:.3f}").font = NORMAL_FONT
            col += 1
        ws.cell(row=row, column=col, value=round(md["Avg_R2"], 3)).font = NORMAL_FONT
        row += 1

    # MolDualNet
    moldualnet_r2 = {"ESOL": 0.918, "FreeSolv": 0.945, "Lipophilicity": 0.768, "BACE": 0.705}
    bold = Font(name="Arial", bold=True, size=10)
    ws.cell(row=row, column=1, value="MolDualNet (ours)").font = bold
    ws.cell(row=row, column=2, value="Multi").font = bold
    col = 3
    for task in ["ESOL", "FreeSolv", "Lipophilicity", "BACE"]:
        ws.cell(row=row, column=col, value=f"{moldualnet_r2[task]:.3f}").font = bold
        col += 1
    ws.cell(row=row, column=col, value="0.834").font = bold

    row += 2
    add_note(ws, row, "All baselines trained per-task on identical Bemis-Murcko scaffold splits with seeds 42, 123, 456.")
    row += 1
    add_note(ws, row, "This table shares source data with Figure 3.")

    auto_width(ws)
    wb.save(os.path.join(OUT_DIR, "Source_Data_Table3.xlsx"))
    print("  Table 3 done")


# ============================================================
# Table 4 — Ablation study
# ============================================================
def generate_table4():
    wb = Workbook()
    ws = wb.active
    ws.title = "Ablation Study"

    with open(os.path.join(SUPP_DATA, "ablation_statistics.json"), "r") as f:
        data = json.load(f)

    configs = [
        ("Full model", "full"),
        ("− Cross-attention", "no_cross_attention"),
        ("− Expert features", "no_expert"),
        ("− 3D geometry", "no_3d"),
        ("Graph only", "gnn_only"),
    ]
    tasks = ["ESOL_logS", "Lipophilicity_logD", "FreeSolv_hydration", "BACE_pIC50"]
    task_labels = ["ESOL", "Lipophilicity", "FreeSolv", "BACE"]

    # Sheet 1: Summary (mean ± std)
    headers = ["Configuration"] + [f"{t} R² (mean ± std)" for t in task_labels] + ["Avg R²", "Δ vs Full"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    full_avg = None
    row = 2
    for label, key in configs:
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        task_means = []
        for col_idx, task in enumerate(tasks, 2):
            td = data["table"][key][task]
            ws.cell(row=row, column=col_idx,
                    value=f"{td['mean']:.3f} ± {td['std']:.3f}").font = NORMAL_FONT
            task_means.append(td["mean"])
        avg = sum(task_means) / len(task_means)
        ws.cell(row=row, column=6, value=round(avg, 3)).font = NORMAL_FONT
        if key == "full":
            full_avg = avg
            ws.cell(row=row, column=7, value="0.000").font = NORMAL_FONT
        else:
            ws.cell(row=row, column=7, value=f"{avg - full_avg:+.3f}").font = NORMAL_FONT
        row += 1

    # Sheet 2: Individual seed values
    ws2 = wb.create_sheet(title="Per-Seed Values")
    headers2 = ["Configuration", "Task", "Seed 42", "Seed 123", "Seed 456", "Seed 789", "Seed 1024"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=i, value=h)
    style_header(ws2, 1, len(headers2))

    row = 2
    for label, key in configs:
        for task, task_label in zip(tasks, task_labels):
            ws2.cell(row=row, column=1, value=label).font = NORMAL_FONT
            ws2.cell(row=row, column=2, value=task_label).font = NORMAL_FONT
            vals = data["table"][key][task]["values"]
            for i, v in enumerate(vals):
                ws2.cell(row=row, column=3 + i, value=round(v, 4)).font = NORMAL_FONT
            row += 1

    # Sheet 3: Statistical significance
    ws3 = wb.create_sheet(title="Statistical Tests")
    headers3 = ["Comparison", "t-statistic", "p-value"]
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=i, value=h)
    style_header(ws3, 1, len(headers3))

    row = 2
    for comp, vals in data["significance"].items():
        ws3.cell(row=row, column=1, value=comp.replace("_", " ")).font = NORMAL_FONT
        ws3.cell(row=row, column=2, value=round(vals["t_stat"], 4)).font = NORMAL_FONT
        ws3.cell(row=row, column=3, value=round(vals["p_value"], 6)).font = NORMAL_FONT
        row += 1

    row += 1
    add_note(ws3, row, f"Seeds: {data['seeds']}, n_seeds: {data['n_seeds']}, epochs: {data['epochs']}")

    auto_width(ws)
    auto_width(ws2)
    auto_width(ws3)
    wb.save(os.path.join(OUT_DIR, "Source_Data_Table4.xlsx"))
    print("  Table 4 done")


# ============================================================
# Table 5 — Cross-dataset generalization
# ============================================================
def generate_table5():
    wb = Workbook()
    ws = wb.active
    ws.title = "Generalization Metrics"

    with open(os.path.join(SUPP_DATA, "cross_dataset_validation_results.json"), "r") as f:
        data = json.load(f)

    headers = ["Task", "Source", "n", "R²", "R² 95% CI low", "R² 95% CI high",
               "RMSE", "RMSE 95% CI low", "RMSE 95% CI high",
               "MAE", "MAE 95% CI low", "MAE 95% CI high",
               "Pearson r", "Spearman ρ", "Mean Signed Error"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    for task in ["FreeSolv_hydration", "Lipophilicity_logD", "ESOL_logS", "BACE_pIC50"]:
        m = data["moldualnet_metrics"][task]
        ws.cell(row=row, column=1, value=task).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=data["validation_types"][task]).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=m["n"]).font = NORMAL_FONT
        ws.cell(row=row, column=4, value=m["R2"]["value"]).font = NORMAL_FONT
        ws.cell(row=row, column=5, value=m["R2"]["CI_95_lo"]).font = NORMAL_FONT
        ws.cell(row=row, column=6, value=m["R2"]["CI_95_hi"]).font = NORMAL_FONT
        ws.cell(row=row, column=7, value=m["RMSE"]["value"]).font = NORMAL_FONT
        ws.cell(row=row, column=8, value=m["RMSE"]["CI_95_lo"]).font = NORMAL_FONT
        ws.cell(row=row, column=9, value=m["RMSE"]["CI_95_hi"]).font = NORMAL_FONT
        ws.cell(row=row, column=10, value=m["MAE"]["value"]).font = NORMAL_FONT
        ws.cell(row=row, column=11, value=m["MAE"]["CI_95_lo"]).font = NORMAL_FONT
        ws.cell(row=row, column=12, value=m["MAE"]["CI_95_hi"]).font = NORMAL_FONT
        ws.cell(row=row, column=13, value=m["Pearson_r"]["value"]).font = NORMAL_FONT
        ws.cell(row=row, column=14, value=m["Spearman_rho"]["value"]).font = NORMAL_FONT
        ws.cell(row=row, column=15, value=round(m["Mean_Signed_Error"], 4)).font = NORMAL_FONT
        row += 1

    auto_width(ws)
    wb.save(os.path.join(OUT_DIR, "Source_Data_Table5.xlsx"))
    print("  Table 5 done")


# ============================================================
# Table 6 — Baseline comparison on shifted/external sets
# ============================================================
def generate_table6():
    wb = Workbook()
    ws = wb.active
    ws.title = "Baseline Comparison"

    with open(os.path.join(SUPP_DATA, "cross_dataset_validation_results.json"), "r") as f:
        cross_data = json.load(f)
    with open(os.path.join(SUPP_DATA, "bace_baseline_comparison.json"), "r") as f:
        bace_data = json.load(f)

    headers = ["Task", "Method", "n", "RMSE", "R²", "Pearson r"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    rows_data = [
        # ESOL
        ("ESOL", "MolDualNet", 8192, 1.637, 0.542, 0.763),
        ("ESOL", "Delaney equation", 8192, 2.622, -0.177, 0.630),
        # Lipophilicity
        ("Lipophilicity", "MolDualNet", 630, 0.511, 0.808, 0.901),
        ("Lipophilicity", "RDKit Crippen log P", 630, 1.762, -1.280, 0.393),
        # BACE
        ("BACE", "MolDualNet", 7722, 1.103, 0.261, 0.540),
        ("BACE", "Random Forest (Morgan FP)", 7722,
         bace_data["baselines"]["RandomForest_MorganFP"]["rmse"],
         round(bace_data["baselines"]["RandomForest_MorganFP"]["r2"], 3),
         round(bace_data["baselines"]["RandomForest_MorganFP"]["pearson_r"], 3)),
        ("BACE", "Gradient Boosted Trees (Morgan FP)", 7722,
         round(bace_data["baselines"]["GBT_MorganFP"]["rmse"], 3),
         round(bace_data["baselines"]["GBT_MorganFP"]["r2"], 3),
         round(bace_data["baselines"]["GBT_MorganFP"]["pearson_r"], 3)),
    ]

    for r, rd in enumerate(rows_data, 2):
        for c, val in enumerate(rd, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = NORMAL_FONT
            if isinstance(val, str) and "MolDualNet" in val:
                cell.font = Font(name="Arial", bold=True, size=10)

    row = len(rows_data) + 3
    add_note(ws, row, "Baselines evaluated on the same shifted or external test sets as MolDualNet.")
    row += 1
    add_note(ws, row, "Delaney equation: linear ESOL model (Delaney, JCICS 2004). "
             "RDKit Crippen log P: Wildman-Crippen partition coefficient (Wildman & Crippen, JCICS 1999).")

    auto_width(ws)
    wb.save(os.path.join(OUT_DIR, "Source_Data_Table6.xlsx"))
    print("  Table 6 done")


# ============================================================
# Main
# ============================================================
if __name__ == "__main__":
    print(f"Generating Source Data Excel files → {OUT_DIR}/")
    print()

    print("Figures:")
    generate_fig1()
    generate_fig2()
    generate_fig3()
    generate_fig4()
    generate_fig5()
    generate_fig6()

    print()
    print("Tables:")
    generate_table1()
    generate_table2()
    generate_table3()
    generate_table4()
    generate_table5()
    generate_table6()

    print()
    print(f"Done! {len(os.listdir(OUT_DIR))} files generated in {OUT_DIR}/")
